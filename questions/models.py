import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, NamedStyle, Border, Side
from io import BytesIO
from django.core.files import File
from django.db.models import QuerySet
from django.db import models
from users.models import User
from telegram import Update
from telegram.ext import CallbackContext
from utils.models import CreateTracker, datetime_str
from typing import Tuple
from tgbot.handlers.admin.static_text import no_questions


QUESTION_MAX_LENGTH = 100


class Question(CreateTracker):
    msg_id = models.BigAutoField(
        primary_key=True, verbose_name="Номер сообщения"
    )  # telegram_id
    user = models.ForeignKey(
        User, verbose_name="Отправитель", on_delete=models.DO_NOTHING
    )
    text = models.TextField(verbose_name="Вопрос")

    def __str__(self):
        return f"msg {self.msg_id} from {self.user}"

    @classmethod
    def add_question(cls, update: Update, context: CallbackContext):
        user, created = User.get_user_and_created(update=update, context=context)
        new_question, created = cls.objects.update_or_create(
            msg_id=update.message.message_id, user=user, text=update.message.text
        )
        return new_question, created

    @classmethod
    def remove_question(cls, msg_id=None) -> int:
        if msg_id:
            question_for_remove = cls.objects.get(msg_id)
        else:
            question_for_remove = cls.objects.all()
        count = question_for_remove.count()
        question_for_remove.delete()
        return count

    @classmethod
    def questions_meta(cls, queryset: QuerySet) -> Tuple[int, str, str]:
        count = str(queryset.count())
        if count == "0":
            first_date, last_date = "", ""
        else:
            first_date, last_date = datetime_str(
                queryset.first().created_at
            ), datetime_str(queryset.last().created_at)
        return count, first_date, last_date

    @classmethod
    def generate_excel_file_name(cls, queryset: QuerySet):
        # count = str(queryset.count())
        # first_date, last_date = datetime_str(queryset.first().created_at), datetime_str(
        #     queryset.last().created_at
        # )
        count, first_date, last_date = cls.questions_meta(queryset)
        if count == 0:
            excel_file_name = "На "
        if count.endswith("1"):
            question_numerals = "вопрос"
        elif count.endswith(
            (
                "2",
                "3",
                "4",
            )
        ):
            question_numerals = "вопроса"
        else:
            question_numerals = "вопросов"
        excel_file_name = f"{count} {question_numerals} заданных с {first_date} по {last_date} по Москве"
        return excel_file_name

    def wrap_text(text: str, max_width: int) -> str:
        words = text.split(" ")
        lines = []
        current_line = []

        for word in words:
            if len(" ".join(current_line + [word])) <= max_width:
                current_line.append(word)
            else:
                lines.append(" ".join(current_line))
                current_line = [word]

        if current_line:
            lines.append(" ".join(current_line))

        return "\n".join(lines)

    @classmethod
    def export_question_to_excel(cls) -> Tuple[str, File, int, str, str]:
        questions = cls.objects.all()
        questions_count, first_date, last_date = cls.questions_meta(questions)
        if questions_count == "0":
            excel_file_name = no_questions
            return excel_file_name, None, questions_count, first_date, last_date
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = cls.generate_excel_file_name(questions)
        # Определение стилей
        bold_font = Font(bold=True, size=16)
        big_font = Font(bold=False, size=16)
        centered_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Заголовки таблицы
        headers = ["Номер сообщения", "Отправитель", "", "", "Вопрос", "Задан"]

        for col_num, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.value = header
            cell.font = bold_font
            cell.alignment = centered_alignment
            cell.border = thin_border

        sender_headers = ["@никнейм", "имя", "фамилия"]
        for col_num, header in enumerate(sender_headers, 2):
            cell = worksheet.cell(row=2, column=col_num)
            cell.value = header
            cell.font = bold_font
            cell.alignment = centered_alignment
            cell.border = thin_border

        # Объединение ячеек для двухуровневого заголовка
        worksheet.merge_cells(start_row=1, start_column=2, end_row=1, end_column=4)
        worksheet.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
        worksheet.merge_cells(start_row=1, start_column=5, end_row=2, end_column=5)
        worksheet.merge_cells(start_row=1, start_column=6, end_row=2, end_column=6)

        # Заполнение данными
        for row_num, question in enumerate(questions, 3):
            # Номер сообщения
            cell = worksheet.cell(row=row_num, column=1)
            cell.font = big_font
            cell.value = question.msg_id
            cell.alignment = centered_alignment
            cell.border = thin_border

            # Отправитель
            sender_values = [
                f"@{str(question.user.username)}",
                str(question.user.first_name),
                str(question.user.last_name),
            ]
            for col_num, value in enumerate(sender_values, 2):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.font = big_font
                cell.value = value
                cell.alignment = centered_alignment
                cell.border = thin_border

            # Вопрос
            cell = worksheet.cell(row=row_num, column=5)
            cell.font = big_font
            cell.value = cls.wrap_text(question.text, QUESTION_MAX_LENGTH)
            cell.alignment = Alignment(
                horizontal="left", vertical="center", wrap_text=True
            )  # Добавьте wrap_text=True
            cell.border = thin_border

            # Установка высоты строки в зависимости от количества строк текста вопроса
            wrapped_lines = cell.value.count("\n") + 1
            worksheet.row_dimensions[row_num].height = (
                wrapped_lines * 18
            )  # Выберите подходящий множитель для высоты строки

            # Дата и время
            cell = worksheet.cell(row=row_num, column=6)
            cell.font = big_font
            cell.value = f"{datetime_str(question.created_at)} по Москве"
            cell.alignment = centered_alignment
            cell.border = thin_border

        # Автоматическое изменение ширины столбцов
        for col_index, column_cells in enumerate(worksheet.columns):
            non_merged_cells = [
                cell
                for cell in column_cells
                if not isinstance(cell, openpyxl.cell.MergedCell)
            ]
            if not non_merged_cells:
                continue

            if col_index == 4:  # Индекс столбца "Вопрос" (начиная с 0)
                max_char_width = (
                    max(len(str(cell.value)) for cell in non_merged_cells)
                    / QUESTION_MAX_LENGTH
                )
                target_width = QUESTION_MAX_LENGTH / 2 * max_char_width
            else:
                target_width = max(
                    len(str(cell.value)) + 8 for cell in non_merged_cells
                )

            worksheet.column_dimensions[
                non_merged_cells[0].column_letter
            ].width = target_width

        excel_file = BytesIO()
        workbook.save(excel_file)
        excel_file.seek(0)
        excel_file_name = f"{worksheet.title}.xlsx"
        return excel_file_name, excel_file, questions_count, first_date, last_date
